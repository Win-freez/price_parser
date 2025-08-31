import asyncio
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font

from app.parsers.positiv.positiv import PositiveParserAPI

BASE_DIR = Path(__file__).parent.parent.parent.parent


async def save_products_to_excel(parser: PositiveParserAPI, filename: str):
    wb = Workbook()
    wb.remove(wb.active)  # удаляем стандартный пустой лист

    headers = [
        "Категория", "public_id", "Имя", "Ссылка на картинку", "Код",
        "Описание", "Остатки", "Ед.Измерения",
        "Доступность", "Цена", "Валюта",
        "Ссылка 1С", "Дата публикации", "Дата снятия публикации", "Дата создания"
    ]

    categories = await parser.get_categories()
    categories_with_children = parser.make_categories_with_children(categories)
    main_categories = (c for c in categories_with_children.values() if c.parent_id is None)

    async def process_category(main_category):
        ws = wb.create_sheet(title=main_category.name[:31])  # имя листа не длиннее 31 символа
        ws.append(headers)

        async for depth, category_name, products in parser.walk_categories(main_category):
            indent = "    " * depth
            ws.append([f"{indent}{category_name}"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

            if products:
                for product in products:
                    row = [
                        category_name,
                        product.public_id,
                        product.name,
                        str(product.imageUrl) if product.imageUrl else "",
                        product.code,
                        product.description,
                        product.count,
                        product.unitOfMeasurement,
                        product.isAvailable,
                        product.price,
                        product.currency,
                        product.links1c,
                        product.publishedDate.strftime("%d-%m-%Y") if product.publishedDate else "",
                        product.unpublishedDate.strftime("%d-%m-%Y") if product.unpublishedDate else "",
                        product.createdAt.strftime("%d-%m-%Y") if product.createdAt else "",
                    ]
                    ws.append(row)

                print(f"Записаны {len(products)} товаров категории {category_name} на листе {main_category.name}")

    await asyncio.gather(*(process_category(c) for c in main_categories))

    wb.save(BASE_DIR / filename)
    wb.close()
    print(f"Excel сохранён в {filename}")


async def main():
    async with httpx.AsyncClient() as client:
        parser = PositiveParserAPI(client=client, max_concurrent=100)
        await save_products_to_excel(parser, filename="positiv_products.xlsx")


asyncio.run(main())
